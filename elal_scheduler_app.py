import io
import re
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="מערכת שיבוץ טיסות", page_icon="✈️", layout="wide")

st.markdown("""
<style>
.block-container { padding-top: 1rem; padding-bottom: 2rem; }
.altea-header {
    background: linear-gradient(90deg, #071b3a 0%, #0b3d78 55%, #071b3a 100%);
    color: white;
    border-radius: 14px;
    padding: 18px 22px;
    margin-bottom: 16px;
    box-shadow: 0 8px 24px rgba(7, 27, 58, 0.28);
}
.altea-title { font-size: 30px; font-weight: 900; margin: 0; }
.altea-subtitle { color: #d8e8ff; font-size: 14px; margin-top: 4px; }
.flight-card {
    border: 1px solid #d9e2ef;
    border-radius: 14px;
    margin-bottom: 14px;
    background: #ffffff;
    box-shadow: 0 2px 9px rgba(0,0,0,.05);
    overflow: hidden;
}
.flight-head {
    background: #071b3a;
    color: white;
    padding: 10px 14px;
    display: flex;
    justify-content: space-between;
    gap: 10px;
    flex-wrap: wrap;
}
.flight-main { font-size: 18px; font-weight: 900; }
.flight-meta { color: #cfe2ff; font-size: 13px; font-weight: 700; }
.role-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 12px;
    padding: 12px;
}
.role-panel {
    border: 1px solid #e3e9f3;
    border-radius: 12px;
    background: #f8fbff;
    padding: 10px 12px;
}
.role-panel-title {
    font-size: 14px;
    font-weight: 900;
    color: #0b3d78;
    margin-bottom: 8px;
    border-bottom: 1px solid #d8e5f7;
    padding-bottom: 5px;
}
.line {
    font-size: 14px;
    padding: 5px 7px;
    margin-bottom: 5px;
    border-radius: 8px;
    background: white;
    border: 1px solid #edf1f7;
}
.teamlead { color: #5e168a; font-weight: 900; border-right: 5px solid #8e24aa; }
.inspector { color: #a30d19; font-weight: 900; border-right: 5px solid #d32f2f; }
.guard { color: #106b2f; font-weight: 900; border-right: 5px solid #2e7d32; }
.missing { color: #9b0000; font-weight: 900; background: #ffe8e8; border: 1px solid #ffb8b8; }
.agent { color: #1f2933; border-right: 5px solid #9fb7d7; }
@media (max-width: 800px) {
    .role-grid { grid-template-columns: 1fr; }
    .altea-title { font-size: 23px; }
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="altea-header">
  <div class="altea-title">✈️ ALTEA-STYLE FLIGHT STAFFING CONTROL</div>
  <div class="altea-subtitle">מערכת שיבוץ מבצעית | חוקים מלאים | TSA | טרייני | משמרות | חפיפות | הפסקות | עומס</div>
</div>
""", unsafe_allow_html=True)


USA_TSA_DESTS = {"JFK", "LAX", "EWR", "FLL", "MIA", "BOS"}
QUEUE_DESTS = {"CDG", "LHR", "JFK", "EWR", "BCN"}
TWO_TEAM_LEADS_DESTS = {"BKK", "HKT"}

NARROW_REG_PREFIXES = ("EH", "EK")
WIDE_REG_PREFIXES = ("EC", "ED", "ER")
REMOTE_GATES = {"D1", "D1A", "C1", "C1A", "B1", "B1A", "E1", "E1A"}

ROLE_COLUMNS = [
    "ראש צוות", "דייל", "מתאם תורים", "מפקח TSA", "שומר TSA",
    "חונך רצים", "מסמיך רצים", "טרייני רצ",
]

ROLE_ORDER = ["ראש צוות", "טרייני רצ", "דייל", "מתאם תורים", "מפקח TSA", "שומר TSA"]


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "nat"} else text


def normalize_yes_no(value) -> str:
    text = clean_text(value).upper()
    return "כן" if text in {"Y", "YES", "כן", "TRUE", "1", "V", "✓", "✔"} else "לא"


def parse_times(value):
    text = clean_text(value)
    match = re.search(r"(\d{2}:\d{2})\s*\((\d{2}:\d{2})\)", text)
    if match:
        return match.group(1), match.group(2)
    match = re.search(r"(\d{2}:\d{2})", text)
    if match:
        return match.group(1), ""
    return "", ""


def to_datetime_time(value):
    return datetime.strptime(clean_text(value), "%H:%M")


def time_to_minutes(value):
    h, m = clean_text(value).split(":")
    return int(h) * 60 + int(m)


def minutes_between(start, end):
    return int((end - start).total_seconds() / 60)


def load_daily_schedule(uploaded_file):
    raw = pd.read_excel(uploaded_file, sheet_name="דוח שיבוץ טיסות - המראות")
    raw.columns = raw.columns.astype(str).str.strip()

    flights = []
    for _, row in raw.iterrows():
        flight = row.get("Unnamed: 8")
        time_text = row.get("Unnamed: 7")
        destination = row.get("Unnamed: 6")
        aircraft = row.get("Unnamed: 5")

        if pd.isna(flight):
            continue

        flight_text = clean_text(flight)
        if not flight_text.startswith("LY"):
            continue

        departure, boarding = parse_times(time_text)

        flights.append({
            "טיסה": flight_text,
            "יעד": clean_text(destination).upper(),
            "המראה": departure,
            "בורדינג": boarding,
            "גייט": "",
            "סוג מטוס": clean_text(aircraft),
            "רישוי": "",
            "נוסעים": "",
            "טרייני רצ": "לא",
            "סוג הכשרה": "",
        })

    return pd.DataFrame(flights)


def normalize_employees(df):
    df = df.copy()
    df.columns = df.columns.astype(str).str.strip()

    if "שם" not in df.columns:
        raise ValueError("בקובץ העובדים חייבת להיות עמודה בשם: שם")

    aliases = {
        "מפקח tsa": "מפקח TSA",
        "שומר tsa": "שומר TSA",
        "טרייני ר״צ": "טרייני רצ",
        'טרייני ר"צ': "טרייני רצ",
        "ראש צוות חונך": "חונך רצים",
        "ראש צוות מסמיך": "מסמיך רצים",
    }

    for old, new in aliases.items():
        if old in df.columns and new not in df.columns:
            df[new] = df[old]

    df["שם"] = df["שם"].apply(clean_text)
    df = df[df["שם"] != ""].copy()

    for col in ROLE_COLUMNS:
        if col not in df.columns:
            df[col] = "לא"
        df[col] = df[col].apply(normalize_yes_no)

    for col in ["תחילת משמרת", "סוף משמרת"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].apply(clean_text)

    return df


def get_body_type(flight):
    reg = clean_text(flight.get("רישוי", "")).upper()
    aircraft = clean_text(flight.get("סוג מטוס", "")).upper()

    if reg.startswith(NARROW_REG_PREFIXES):
        return "צר גוף"
    if reg.startswith(WIDE_REG_PREFIXES):
        return "רחב גוף"
    if aircraft.startswith(("737", "738", "739", "E")):
        return "צר גוף"
    return "רחב גוף"


def is_remote_gate(gate):
    return clean_text(gate).upper() in REMOTE_GATES


def get_pax(flight):
    value = flight.get("נוסעים", 0)
    if pd.isna(value) or clean_text(value) == "":
        return 0
    try:
        return int(float(value))
    except Exception:
        return 0


def get_requirements(flight):
    dest = clean_text(flight["יעד"]).upper()
    body = get_body_type(flight)
    pax = get_pax(flight)

    req = {
        "ראש צוות": 1,
        "דייל": 1,
        "מתאם תורים": 0,
        "מפקח TSA": 0,
        "שומר TSA": 0,
        "טרייני רצ": 0,
    }

    if body == "צר גוף":
        req["דייל"] = 1 if pax <= 150 else 2
    else:
        req["דייל"] = 5 if dest in USA_TSA_DESTS else 3

    if dest in TWO_TEAM_LEADS_DESTS:
        req["ראש צוות"] = 2

    if dest in QUEUE_DESTS:
        req["מתאם תורים"] = 1

    if dest in USA_TSA_DESTS:
        req["מפקח TSA"] = 1
        req["שומר TSA"] = 1

    if clean_text(flight.get("טרייני רצ", "")) == "כן":
        req["טרייני רצ"] = 1

    return req


def role_start_time(flight, role):
    departure = to_datetime_time(flight["המראה"])
    body = get_body_type(flight)
    remote_narrow = body == "צר גוף" and is_remote_gate(flight.get("גייט", ""))

    if role in {"ראש צוות", "מתאם תורים", "טרייני רצ"}:
        minutes_before = 60 if body == "צר גוף" else 75
    elif role == "דייל":
        minutes_before = 50 if body == "צר גוף" else 65
    elif role in {"מפקח TSA", "שומר TSA"}:
        minutes_before = 120
    else:
        minutes_before = 60

    if remote_narrow:
        minutes_before += 10

    return departure - timedelta(minutes=minutes_before)


def role_end_time(flight):
    return to_datetime_time(flight["המראה"])


def shift_length(emp):
    shift_start = clean_text(emp.get("תחילת משמרת", ""))
    shift_end = clean_text(emp.get("סוף משמרת", ""))

    if shift_start == "" or shift_end == "":
        return 0

    s = time_to_minutes(shift_start)
    e = time_to_minutes(shift_end)

    if e < s:
        e += 24 * 60

    return e - s


def required_break(emp):
    length = shift_length(emp)
    if length > 9 * 60 + 30:
        return 65
    if length > 6 * 60:
        return 45
    return 0


def is_within_shift(emp, task_start, task_end):
    shift_start = clean_text(emp.get("תחילת משמרת", ""))
    shift_end = clean_text(emp.get("סוף משמרת", ""))

    if shift_start == "" or shift_end == "":
        return True

    task_start_min = task_start.hour * 60 + task_start.minute
    task_end_min = task_end.hour * 60 + task_end.minute

    s = time_to_minutes(shift_start)
    e = time_to_minutes(shift_end)

    if e < s:
        if task_start_min < s:
            task_start_min += 24 * 60
            task_end_min += 24 * 60
        e += 24 * 60

    return task_start_min >= s and task_end_min <= e


def assigned_minutes(assignments, emp_name):
    total = 0

    for task in assignments:
        if task["עובד"] != emp_name:
            continue
        if clean_text(task.get("התחלה", "")) == "" or clean_text(task.get("סיום", "")) == "":
            continue

        total += minutes_between(
            to_datetime_time(task["התחלה"]),
            to_datetime_time(task["סיום"]),
        )

    return total


def has_room_for_break(assignments, emp, emp_name, start, end):
    length = shift_length(emp)
    if length == 0:
        return True

    current = assigned_minutes(assignments, emp_name)
    new = minutes_between(start, end)

    return current + new + required_break(emp) <= length


def is_available(assignments, emp_name, start, end):
    buffer = timedelta(minutes=5)

    for task in assignments:
        if task["עובד"] != emp_name:
            continue
        if clean_text(task.get("התחלה", "")) == "" or clean_text(task.get("סיום", "")) == "":
            continue

        existing_start = to_datetime_time(task["התחלה"])
        existing_end = to_datetime_time(task["סיום"])

        if not (start >= existing_end + buffer or end <= existing_start - buffer):
            return False

    return True


def count_all_tasks(assignments, emp_name):
    return sum(1 for task in assignments if task["עובד"] == emp_name)


def count_team_lead_tasks(assignments, emp_name):
    return sum(
        1 for task in assignments
        if task["עובד"] == emp_name and str(task["תפקיד"]).startswith("ראש צוות")
    )


def sort_candidates(candidates, assignments, role):
    candidates = candidates.copy()

    if candidates.empty:
        return candidates

    if role == "ראש צוות":
        candidates["_score"] = candidates["שם"].apply(lambda name: count_team_lead_tasks(assignments, name))
        return candidates.sort_values("_score")

    if role == "דייל":
        candidates["_score1"] = candidates.apply(
            lambda row: 0 if str(row.get("ראש צוות", "")).strip() == "כן" else 1,
            axis=1,
        )
        candidates["_score2"] = candidates["שם"].apply(lambda name: count_all_tasks(assignments, name))
        return candidates.sort_values(["_score1", "_score2"])

    candidates["_score"] = candidates["שם"].apply(lambda name: count_all_tasks(assignments, name))
    return candidates.sort_values("_score")


def has_required_mentor(assignments_for_flight, employees_df, training_type):
    required_col = "מסמיך רצים" if training_type == "הסמכה" else "חונך רצים"

    for task in assignments_for_flight:
        if str(task["תפקיד"]).startswith("ראש צוות") and "❌" not in str(task["עובד"]):
            emp = employees_df[employees_df["שם"] == task["עובד"]]
            if not emp.empty and str(emp.iloc[0].get(required_col, "")).strip() == "כן":
                return True

    return False


def build_schedule(flights_df, employees_df):
    assignments = []
    role_order = ROLE_ORDER
    flights_df = flights_df[flights_df["המראה"].astype(str).str.strip() != ""].copy()

    for _, flight in flights_df.iterrows():
        req = get_requirements(flight)
        used_on_flight = set()
        assignments_for_flight = []

        trainee_needed = req.get("טרייני רצ", 0) > 0
        training_type = clean_text(flight.get("סוג הכשרה", "חניכה")) or "חניכה"

        for role in role_order:
            amount = req.get(role, 0)

            for i in range(amount):
                start = role_start_time(flight, role)
                end = role_end_time(flight)

                if role not in employees_df.columns:
                    employees_df[role] = "לא"

                candidates = employees_df[
                    (employees_df[role].astype(str).str.strip() == "כן") &
                    (~employees_df["שם"].isin(used_on_flight))
                ].copy()

                if role == "ראש צוות" and trainee_needed:
                    mentor_col = "מסמיך רצים" if training_type == "הסמכה" else "חונך רצים"
                    if mentor_col not in candidates.columns:
                        candidates[mentor_col] = "לא"

                    mentors = candidates[candidates[mentor_col].astype(str).str.strip() == "כן"]

                    if not mentors.empty:
                        candidates = mentors

                candidates = sort_candidates(candidates, assignments, role)

                selected = None

                for _, emp in candidates.iterrows():
                    name = emp["שם"]

                    if (
                        is_within_shift(emp, start, end)
                        and is_available(assignments, name, start, end)
                        and has_room_for_break(assignments, emp, name, start, end)
                    ):
                        selected = name
                        break

                if selected:
                    worker = selected
                    used_on_flight.add(worker)
                else:
                    worker = f"❌ חסר {role}"

                task = {
                    "טיסה": flight["טיסה"],
                    "יעד": flight["יעד"],
                    "תפקיד": role if amount == 1 else f"{role} {i+1}",
                    "עובד": worker,
                    "התחלה": start.strftime("%H:%M"),
                    "סיום": end.strftime("%H:%M"),
                }

                assignments.append(task)
                assignments_for_flight.append(task)

        if trainee_needed and not has_required_mentor(assignments_for_flight, employees_df, training_type):
            assignments.append({
                "טיסה": flight["טיסה"],
                "יעד": flight["יעד"],
                "תפקיד": "בדיקת טרייני",
                "עובד": "❌ אין חונך/מסמיך מתאים בטיסה",
                "התחלה": "",
                "סיום": "",
            })

    return pd.DataFrame(assignments)


def get_employee_shift(employees_df, emp_name):
    emp = employees_df[employees_df["שם"] == emp_name]

    if emp.empty:
        return ""

    emp = emp.iloc[0]
    s = clean_text(emp.get("תחילת משמרת", ""))
    e = clean_text(emp.get("סוף משמרת", ""))

    if s and e:
        return f"{s}-{e}"

    return ""


def build_next_task_labels(result_df, employees_df):
    df = result_df.copy()
    df["טקסט עובד"] = df["עובד"].astype(str)

    timed_df = df[df["התחלה"].astype(str).str.strip() != ""].copy()
    timed_df["_start_dt"] = pd.to_datetime(timed_df["התחלה"], format="%H:%M", errors="coerce")

    break_used = {}

    for idx, row in timed_df.sort_values("_start_dt").iterrows():
        emp = str(row["עובד"]).strip()

        if "❌" in emp:
            df.loc[idx, "טקסט עובד"] = emp
            continue

        emp_row = employees_df[employees_df["שם"] == emp]
        needs_break = not emp_row.empty and required_break(emp_row.iloc[0]) > 0

        future = timed_df[
            (timed_df["עובד"].astype(str) == emp) &
            (timed_df["_start_dt"] > row["_start_dt"])
        ].sort_values("_start_dt")

        if needs_break and not break_used.get(emp, False):
            break_used[emp] = True
            next_text = "הפסקה וחזרה" if future.empty else f'הפסקה וטיסה {future.iloc[0]["טיסה"]}'
        else:
            next_text = "סוף משמרת" if future.empty else f'טיסה {future.iloc[0]["טיסה"]}'

        shift = get_employee_shift(employees_df, emp)
        df.loc[idx, "טקסט עובד"] = f"{emp} ({shift}) - {next_text}" if shift else f"{emp} - {next_text}"

    return df


def build_workload(result_df, employees_df):
    rows = []
    timed = result_df[result_df["התחלה"].astype(str).str.strip() != ""].copy()

    real_workers = sorted([
        worker for worker in timed["עובד"].dropna().unique()
        if "❌" not in str(worker)
    ])

    for emp in real_workers:
        tasks = timed[timed["עובד"] == emp]
        total = 0

        for _, task in tasks.iterrows():
            total += minutes_between(
                to_datetime_time(task["התחלה"]),
                to_datetime_time(task["סיום"]),
            )

        emp_row = employees_df[employees_df["שם"] == emp]
        break_min = required_break(emp_row.iloc[0]) if not emp_row.empty else 0

        rows.append({
            "עובד": emp,
            "משימות": len(tasks),
            "דקות עבודה": total,
            "הפסקה נדרשת": break_min,
            "סה״כ כולל הפסקות": total + break_min,
        })

    return pd.DataFrame(rows)


def build_output_table(flights_df, result_labeled):
    rows = []

    for _, flight in flights_df.iterrows():
        fnum = str(flight["טיסה"]).strip()
        dep = clean_text(flight.get("המראה", ""))
        boarding = clean_text(flight.get("בורדינג", ""))
        aircraft = clean_text(flight.get("סוג מטוס", ""))
        reg = clean_text(flight.get("רישוי", ""))

        tasks = result_labeled[result_labeled["טיסה"].astype(str).str.strip() == fnum]

        management_lines = []
        agents_lines = []

        for _, task in tasks.iterrows():
            role = str(task["תפקיד"])
            text = f'{role}: {task["טקסט עובד"]}'

            if "ראש צוות" in role or "מתאם" in role or "מפקח TSA" in role:
                management_lines.append(text)
            elif "דייל" in role or "שומר TSA" in role or "בדיקת טרייני" in role:
                agents_lines.append(text)

        rows.append({
            "מספר טיסה": fnum,
            "יעד": clean_text(flight["יעד"]),
            "זמנים": f"{dep} ({boarding})" if boarding else dep,
            "מטוס/רישוי": f"{aircraft}\n{reg}".strip(),
            "ראש צוות / מתאם תורים / מפקח TSA": "\n".join(management_lines),
            "דיילים / שומר TSA": "\n".join(agents_lines),
        })

    return pd.DataFrame(rows)


def render_line(line):
    css = "agent"
    if "❌" in line:
        css = "missing"
    elif "ראש צוות" in line:
        css = "teamlead"
    elif "מפקח TSA" in line:
        css = "inspector"
    elif "שומר TSA" in line:
        css = "guard"

    st.markdown(f'<div class="line {css}">{line}</div>', unsafe_allow_html=True)


def render_flight_card(row):
    aircraft = str(row["מטוס/רישוי"]).replace("\n", " / ")
    left_text = str(row["ראש צוות / מתאם תורים / מפקח TSA"])
    right_text = str(row["דיילים / שומר TSA"])

    st.markdown(f"""
    <div class="flight-card">
      <div class="flight-head">
        <div class="flight-main">✈️ {row['מספר טיסה']} → {row['יעד']}</div>
        <div class="flight-meta">🕒 {row['זמנים']} | 🛩️ {aircraft}</div>
      </div>
      <div class="role-grid">
        <div class="role-panel">
          <div class="role-panel-title">👔 ראש צוות / מתאם / TSA</div>
    """, unsafe_allow_html=True)

    if left_text and left_text != "nan":
        for line in left_text.split("\n"):
            render_line(line)
    else:
        render_line("אין שיבוץ")

    st.markdown("""
        </div>
        <div class="role-panel">
          <div class="role-panel-title">🧍 דיילים / שומר TSA</div>
    """, unsafe_allow_html=True)

    if right_text and right_text != "nan":
        for line in right_text.split("\n"):
            render_line(line)
    else:
        render_line("אין שיבוץ")

    st.markdown("""
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)


def to_excel_bytes(output_df, workload_df, schedule_df):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name="שיבוץ")
        workload_df.to_excel(writer, index=False, sheet_name="עומס")
        schedule_df.to_excel(writer, index=False, sheet_name="פירוט גולמי")

        wb = writer.book

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        purple_fill = PatternFill("solid", fgColor="EADCF8")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        green_fill = PatternFill("solid", fgColor="D9EAD3")
        blue_fill = PatternFill("solid", fgColor="D9EAF7")
        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.sheet_view.rightToLeft = True
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border

            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        ws = wb["שיבוץ"]

        for row in ws.iter_rows(min_row=2):
            col_e = str(row[4].value)
            col_f = str(row[5].value)

            if "❌" in col_e or "מפקח TSA" in col_e:
                row[4].fill = red_fill
            elif "ראש צוות" in col_e:
                row[4].fill = purple_fill

            if "❌" in col_f:
                row[5].fill = red_fill
            elif "שומר TSA" in col_f:
                row[5].fill = green_fill
            elif "דייל" in col_f:
                row[5].fill = blue_fill

        widths = {"A": 16, "B": 10, "C": 18, "D": 18, "E": 60, "F": 60}

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 70

    output.seek(0)
    return output


with st.sidebar:
    st.header("📂 העלאת קבצים")
    daily_file = st.file_uploader("קובץ סידור יומי", type=["xlsx"])
    employees_file = st.file_uploader("קובץ עובדים / הסמכות", type=["xlsx"])

if not daily_file or not employees_file:
    st.info("העלי קובץ סידור יומי וקובץ עובדים כדי להתחיל.")
    st.stop()

try:
    flights_df = load_daily_schedule(daily_file)
    employees_df = pd.read_excel(employees_file)
    employees_df = normalize_employees(employees_df)
except Exception as exc:
    st.error("לא הצלחתי לקרוא את הקבצים.")
    st.exception(exc)
    st.stop()

st.subheader("🛫 טיסות מהסידור היומי")
st.caption("השלימי כאן גייט, רישוי, נוסעים, טרייני וסוג הכשרה. זה משפיע על כל החוקים.")

edited_flights = st.data_editor(
    flights_df,
    use_container_width=True,
    num_rows="dynamic",
    key="flights_editor",
)

m1, m2, m3, m4 = st.columns(4)
m1.metric("טיסות", len(edited_flights))
m2.metric("עובדים", len(employees_df))
m3.metric("יעדי TSA", edited_flights["יעד"].isin(list(USA_TSA_DESTS)).sum())
m4.metric("טיסות עם טרייני", (edited_flights["טרייני רצ"].astype(str).str.strip() == "כן").sum())

if st.button("🚀 בנה שיבוץ", use_container_width=True):
    try:
        schedule_df = build_schedule(edited_flights, employees_df)
        labeled_df = build_next_task_labels(schedule_df, employees_df)
        workload_df = build_workload(schedule_df, employees_df)
        output_df = build_output_table(edited_flights, labeled_df)

        st.success("השיבוץ נבנה בהצלחה.")

        tab_schedule, tab_missing, tab_workload, tab_raw = st.tabs(
            ["✈️ שיבוץ", "❌ חוסרים", "📊 עומס עובדים", "🧾 פירוט גולמי"]
        )

        with tab_schedule:
            st.subheader("✈️ Altea Style Assignment Board")

            search = st.text_input("🔎 חיפוש לפי טיסה / יעד / עובד")
            only_missing = st.checkbox("הצג רק טיסות עם חוסר")

            display_df = output_df.copy()

            if only_missing:
                display_df = display_df[
                    display_df.astype(str).apply(lambda row: row.str.contains("❌", na=False).any(), axis=1)
                ]

            if search:
                mask = display_df.astype(str).apply(
                    lambda row: row.str.contains(search, case=False, na=False).any(),
                    axis=1,
                )
                display_df = display_df[mask]

            for _, row in display_df.iterrows():
                render_flight_card(row)

        with tab_missing:
            st.subheader("❌ חוסרים")
            missing = schedule_df[schedule_df["עובד"].astype(str).str.contains("❌", na=False)]

            if missing.empty:
                st.success("אין חוסרים 🎉")
            else:
                st.warning(f"נמצאו {len(missing)} חוסרים")
                st.dataframe(missing, use_container_width=True)

        with tab_workload:
            st.subheader("📊 עומס עובדים")
            st.dataframe(workload_df, use_container_width=True)

        with tab_raw:
            st.subheader("🧾 פירוט גולמי")
            st.dataframe(labeled_df, use_container_width=True)

        excel_data = to_excel_bytes(output_df, workload_df, labeled_df)

        st.download_button(
            "⬇️ הורדת אקסל מלא",
            data=excel_data,
            file_name="SCHEDULE_ALTEA_STYLE_FULL.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    except Exception as exc:
        st.error("הייתה שגיאה בבניית השיבוץ.")
        st.exception(exc)
